# Find dead links in EarthCube database
from openpyxl import load_workbook, cell
from openpyxl.cell import coordinate_from_string, column_index_from_string
import requests
from datetime import datetime, date, time

# constants
START_ROW = 2 # row to start
END_ROW = 57 # row to end
TITLE_COL = 'A'
LINK_COL = 'B'  # column with URLs
FIRST_CHECK_COL = 'D' # column that holds first checked date
FRIST_CHECK_EARTHCUBE = '2014-05-07 11:25:12' # the first time that the first resource in was checked
LAST_CHECK_COL = 'E'
URL_STATUS_COL = 'F'
STATUS_CHANGED_COL = 'G'
HTTP = 'http://'

#open workbook and get sheet
wb = load_workbook('TDWG databases.xlsx')
ws = wb.get_active_sheet()

   
def timeStamp( cellRow ):
    print 'Applying time stamp...'
    ws['%s%s'%(FIRST_CHECK_COL, cellRow)].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['%s%s'%(LAST_CHECK_COL, cellRow)].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['%s%s'%(URL_STATUS_COL, cellRow)].value = status    
    
print ('Checking links...')
for row in ws.range('%s%s:%s%s'%(LINK_COL, START_ROW, LINK_COL, END_ROW)):
    for cell in row:
        coordinate = coordinate_from_string(cell.get_coordinate())
        title = ws.cell('%s%s'%(TITLE_COL, coordinate[1])).value
        # follow link in cell
        if cell.value and HTTP in cell.value:
            try:
                # see if the link works
                r = requests.get(cell.value, timeout=6)
                c = r.status_code
            except requests.ConnectionError:
                status = 'Connection Error'
                print ('{}: ConnectionError'.format(ws.cell('%s%s'%(TITLE_COL, coordinate[1])).value))
            except requests.HTTPError:
                status = 'HTTP Error'
                print ('{}: HTTP Error'.format(ws.cell('%s%s'%(TITLE_COL, coordinate[1])).value))
            except requests.TooManyRedirects:
                status = 'Too many redirects'
                print ('{}: Too many redirects'.format(ws.cell('%s%s'%(TITLE_COL, coordinate[1])).value))
            except requests.Timeout:
                status = 'Timeout'
                print ('{}: Timeout'.format(ws.cell('%s%s'%(TITLE_COL, coordinate[1])).value))
            else:
                coordinate = coordinate_from_string(cell.get_coordinate())
                if c != 200:
                    status = c
                    print ('{}: bad link'.format(ws.cell('%s%s'%(TITLE_COL, coordinate[1])).value))
                    print ('    {}'.format(c))
                else:
                    status = 'Working'
            #finally:
            #    timeStamp( coordinate[1] )
                
        else:
            coordinate = coordinate_from_string(cell.get_coordinate())
            print ('{}: no link'.format(ws.cell('%s%s'%(TITLE_COL, coordinate[1])).value))
            ws['%s%s'%(URL_STATUS_COL, coordinate[1])].value = 'No URL provided'
            
wb.save('TDWG databases.xlsx')
print ('Done.')