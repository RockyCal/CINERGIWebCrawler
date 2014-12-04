from bs4 import BeautifulSoup
import requests
import queue
from openpyxl import Workbook, cell
from urllib.request import urlopen, Request
import threading
from write import write_resource
import time
from check_link import check_link
from make_headers import make_headers
from build_text import build_text
from build_labels import build_labels
from Resource import Resource
# <editor-fold desc="Protocol constants">
HTTP = 'http://'
preFTP = 'ftp://'
# </editor-fold>


class Thread(threading.Thread):
    def __init__(self, stack, tier, count, delay):
        threading.Thread.__init__(self)
        self.stack = stack
        self.tier = tier
        self.count = count
        self.delay = delay

    def run(self):
        crawl(self.stack, self.tier, self.delay)


class ThreadClass(threading.Thread):
    def __init__(self, que, count, delay):
        threading.Thread.__init__(self)
        self.queue = que
        self.count = count
        self.delay = delay

    def run(self):
        while True:
            items = self.queue.get()
            stack = items[0]
            new_tier = items[1]
            crawl(stack, new_tier, self.delay)
            self.queue.task_done()


# <editor-fold desc="Functions">

"""
name: crawl
stack is stack of links
"""


def crawl(stack, new_tier, delay):
    while len(stack) > 0:
        # Make instance of Resource
        resource = Resource(stack.pop(0))
        if resource.status is "working":
            urlopen(resource.link)
            if resource.link not in visited:
                resource.get_resource_data()
                visited.append(resource.link)
                new_tier.append(resource)
        time.sleep(delay)

# </editor-fold>

"""
Program begins.
"""
# Total visited links
visited = []
# Total working links found
urls = []
# Total broken links


# Create excel file
wb = Workbook()
filename = input("Enter a title for the excel file: ")
#filename = 'GreenSeas_9_15.xlsx'

# <editor-fold desc="Build org, country, social media">
# List of organizations
org_url = 'http://opr.ca.gov/s_listoforganizations.php'
# List of country codes
country_codes_url = 'http://www.thrall.org/domains.htm'
# Check functioning of organizations list
if check_link(org_url) is "working":
    t = requests.get(org_url)
    orgText = t.text
    soupOrg = BeautifulSoup(orgText)
    orgsOfficial = build_labels(soupOrg)

    ws2 = wb.create_sheet()
    ws2.title = 'List of Official Organizations'
    if len(orgsOfficial) > 0:
        start_row = ws2.get_highest_row() + 1
        last_row = (start_row + len(orgsOfficial)) - 1
        t = 0
        for row in ws2.range('%s%s:%s%s' % ('A', start_row, 'A', last_row)):
            for cell in row:
                cell.value = orgsOfficial[t]
                t += 1
else:
    print("Error with org url")

# Check functioning of country codes list
if check_link(country_codes_url) is "working":
    s = requests.get(country_codes_url)
    counText = s.text
    soupCoun = BeautifulSoup(counText)
    countriesOfficial = build_text(soupCoun)
    countriesOfficial.append("EU - European Union")
    # Get rid of first four, random values
    for t in range(0, 4):
        countriesOfficial.pop(0)

    ws3 = wb.create_sheet()
    ws3.title = 'List of Country Codes'
    if len(countriesOfficial) > 0:
        start_row = ws3.get_highest_row() + 1
        last_row = (start_row + len(countriesOfficial)) - 1
        t = 0
        for row in ws3.range('%s%s:%s%s' % ('A', start_row, 'A', last_row)):
            for cell in row:
                cell.value = countriesOfficial[t]
                t += 1
else:
    print("Error with country codes url")

# </editor-fold>

"""
Resources is the list of actual resources
"""
resources = []
"""
Tiers are levels of link crawling. The start url is tier0.
The links on the page of the start url are tier1. All the links
on all the pages of those urls in tier1 are tier2. And so on.
The first link of every tier is the source link
"""
tiers = []
tier0 = []
tier1 = []


all_visited = []

prompt = input("Enter 1 if you have made a text file with each link separated by a newline or a comma. ")
mode = int(prompt)

if mode is 1:
    print("Found list")
    url_list = []
    f = open('ListOfLinks.txt', 'r')
    for link in f:
        url_list.append(link)
        print(link)
        r = Resource(link)
        print(r)
        r.get_resource_data()
        tier0.append(r)
        resources.append(tier0)
        r.find_links()
        links = r.links_found
     #   for each in links:
         #   print(each)
        if links is not None:
            print("Not none")
            for each2 in links:
              #  print(each2)
                t = Resource(each2)
                t.get_resource_data()

                visited.append(each2)
                tier1.append(t)
                resources.append(tier1)

                # Number of threads for second layer (tier1)
                t1_num_threads = 1
                wait = 0

                # Create a size based on length of links found
                chunksize = int((len(links))/t1_num_threads)
                threads = []

                # Pass in chunks of res0.links_found to threads to be processed
                for i in range(t1_num_threads):
                    t = Thread(r.links_found[chunksize*i:(chunksize*(i+1))], tier1, i, wait)
                    threads.append(t)
                    t.start()

                # Wait for all threads to finish
                # Tier1 should be populated with resources
                for t in threads:
                    t.join()

                # process remaining links
                remains = []
                remaining = len(r.links_found) % t1_num_threads
                rem = (len(r.links_found) - remaining)
                for rem in range(len(r.links_found)):
                    remains.append(r.links_found[rem])
                crawl(remains, tier1, wait)

                print("Length tier1: {}".format(len(tier1)))
                # To hold the links found in third layer
                tier2_links = []

                # Find links on the pages in tier1
                for r in tier1:
                    r.find_links()
                    if r.links_found is not None:
                        tier2_links.extend(r.links_found)

                # Add tier1 to resources
                resources.append(tier1)
                tier2 = []

                # Queue for third layer processing
                q = queue.Queue()
                crawl_time = time.clock()
                print("Gathering data from pages...")

                # Create pool of threads, more threads since
                # third layer is larger, naturally
                # Pass queue instance and thread count
                wait = 30
                t2_num_threads = 2
                for j in range(t2_num_threads):
                    thread = ThreadClass(q, j, wait)
                    thread.setDaemon(True)
                    thread.start()

                size = int((len(tier2_links))/t2_num_threads)
                # Populate queue with chunks of links
                # Will be passed into crawl
                for k in range(t2_num_threads):
                    q.put((tier2_links[(size*k):(size*(k+1))], tier2))
                # TODO: fix 429 Too many requests

                # Wait until everything is processed
                q.join()
                resources.append(tier2)
      #  print('Finished crawl. {} process time'.format(time.clock() - crawl_time))
        print("Finished crawl")

#
# for tier in resources:
#     print(len(tier))

# <editor-fold desc="Write to excel">
write_time0 = time.clock()
index = 0
ws = wb.create_sheet(index, str(index))
make_headers(ws)
row = ws.get_highest_row() + 1
term_links = []
for tier in resources:
    for item in tier:
        write_resource(ws, row, item)
        row += 1
    index += 1
    row = 1
    ws = wb.create_sheet(index, str(index))

write_time = time.clock() - write_time0
print('Write time: {}'.format(write_time))
# </editor-fold>

wb.save(filename+".xlsx")