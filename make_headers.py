__author__ = 'Anoushka'
from openpyxl.styles import Style, Font

def make_headers(sheet):
    header_style = Style(font=Font(bold=True))
    sheet.cell('A1').value = 'Title'  # we need to find out how to do
    sheet['A1'].style = header_style
    sheet.cell('B1').value = 'URL'
    sheet['B1'].style = header_style
    sheet.cell('C1').value = 'Organization'
    sheet['C1'].style = header_style
    sheet['D1'].style = header_style
    sheet.cell('D1').value = 'Theme(s)'
    sheet.cell('E1').value = 'Resource Type'
    sheet['E1'].style = header_style
    sheet['F1'].value = "Country"
    sheet['F1'].style = header_style
    sheet['G1'].value = "Social Media?"
    sheet['G1'].style = header_style
    sheet['H1'].value = "Term Definitions"
    sheet['H1'].style = header_style